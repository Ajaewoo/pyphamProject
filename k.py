import openpyxl
from openpyxl import Workbook, load_workbook
#openpyxl는 라이브러리, Workbook은 class
'''
wb=Workbook()
ws=wb.create_sheet('시트 생성')
#시트 만드는 명령어

ws=wb.active
ws['A1']='hello'


wb=load_workbook("코로나.xlsx")
#엑셀 불러오기
ws=wb['선별진료소']
#시트명을 이용해서 불러오기

print(ws['D3'].value)
print(ws.cell(3,2).value)

c=ws['B2:B4']
for i in c:
    for j in i:
        print(j.value)

for i in ws.rows:
    print(i)
#행단위로 출력

print(ws['A'])
#A열의 모든 셀을 출력한다


wb=openpyxl.Workbook
#라이브러리에 객체 생성
ws=wb.active
#워크시트 생성(엑셀 페이지)

wb=openpyxl.load_workbook(filename='stat.xlsx')
#기존 엑셀파일 불러오기
ws=wb.active
#워크시트 생성
ws=wb['Sheet0']
#시트명을 이용해서 불러오기

print(ws['A7'])
#워크시트 'Sheet0'에 A7셀 출력
print(ws['A7'].value)

ws.cell(row=3,column=2)
#3행 2열 출력

wb.save('stat.xlsx')
#파일 저장
'''

wb=openpyxl.Workbook()
ws=wb.active

ws.title='hi'
#엑셀 시트 'hi' 생성

a=['이름','나이']
ws.append(a)
#'이름', '나이' 넣는다 (a1와 a2에 들어간다)

wb.save('hi.xlsx')
#'hi'엑셀을 저장한다


human={
    '홍길동':'20',
    '김길동':'30',
    '이길동':'40',
}
for i in human.items():
    ws.append(i)
#딕셔너리 값 넣기
#위 넣은 값 아래 넣어진다
wb.save('hi.xlsx')
#저장

ws2=wb.create_sheet(tiltle='hello')

ws3=wb.create_sheet(tiltle='bye')
#워크시트 추가로 생성
wb.save('hi.xlsx')
#저장
